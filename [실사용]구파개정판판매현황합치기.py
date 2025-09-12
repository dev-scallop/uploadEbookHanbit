import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
# import pandas as pd  # pandas가 설치되어 있지 않으므로 주석 처리

def merge_sales_with_new_editions(input_file, output_file):
    # 1. 엑셀에서 시트 불러오기 (openpyxl 사용)
    workbook = openpyxl.load_workbook(input_file)
    
    # 도서판매추이 시트 읽기
    sales_sheet = workbook["도서판매추이"]
    sales_data = []
    headers = []
    
    for row in sales_sheet.iter_rows(values_only=True):
        if not headers:
            headers = list(row)
        else:
            sales_data.append(list(row))
    
    # 개정판 정보 시트 읽기
    mapping_sheet = workbook["개정판 정보"]
    mapping_data = []
    mapping_headers = []
    
    for row in mapping_sheet.iter_rows(values_only=True):
        if not mapping_headers:
            mapping_headers = list(row)
        else:
            mapping_data.append(list(row))
    
    # 2. 매핑 dict 생성 (구판 → 개정판)
    y_code_idx = mapping_headers.index('y_code')
    x_code_idx = mapping_headers.index('x_code')
    x_title_idx = mapping_headers.index('x_title')
    
    map_dict = {}
    code_to_title = {}
    for row in mapping_data:
        map_dict[row[y_code_idx]] = row[x_code_idx]
        code_to_title[row[x_code_idx]] = row[x_title_idx]
    
    # 3. "통합코드" 컬럼 생성 (구판이면 개정판으로 치환, 아니면 자기 자신)
    도서코드_idx = headers.index('도서코드')
    통합코드_idx = len(headers)
    통합여부_idx = 통합코드_idx + 1
    
    # 헤더에 새 컬럼 추가
    headers.extend(['통합코드', '통합여부'])
    
    # 데이터에 새 컬럼 추가
    for row in sales_data:
        도서코드 = row[도서코드_idx]
        통합코드 = map_dict.get(도서코드, 도서코드)
        통합여부 = "통합" if 도서코드 in map_dict else "단일"
        row.extend([통합코드, 통합여부])
    
    # 4. 구판+개정판 합산 (일자별 합계 유지)
    # 그룹화를 위한 인덱스 찾기
    도서영역별_idx = headers.index('도서영역별')
    
    # 그룹화 데이터 생성
    grouped_data = {}
    for row in sales_data:
        key = (row[도서영역별_idx], row[통합코드_idx])
        if key not in grouped_data:
            grouped_data[key] = {}
            for i, header in enumerate(headers):
                if header not in ['도서영역별', '도서코드', '도서', '통합코드', '통합여부']:
                    grouped_data[key][i] = 0
        
        for i, header in enumerate(headers):
            if header not in ['도서영역별', '도서코드', '도서', '통합코드', '통합여부']:
                if isinstance(row[i], (int, float)):
                    grouped_data[key][i] += row[i]
    
    # 5. 결과를 새로운 워크북에 저장
    new_workbook = Workbook()
    
    # 원본+통합여부 시트
    ws1 = new_workbook.active
    ws1.title = "원본+통합여부"
    
    # 헤더 추가
    for col, header in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=header)
    
    # 데이터 추가
    for row_idx, row_data in enumerate(sales_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws1.cell(row=row_idx, column=col_idx, value=value)
    
    # 구판+개정판통합 시트
    ws2 = new_workbook.create_sheet("구판+개정판통합")
    
    # 통합된 헤더 생성
    merged_headers = ['도서영역별', '통합코드']
    for header in headers:
        if header not in ['도서영역별', '도서코드', '도서', '통합코드', '통합여부']:
            merged_headers.append(header)
    merged_headers.append('도서명(개정/통합)')
    
    # 헤더 추가
    for col, header in enumerate(merged_headers, 1):
        ws2.cell(row=1, column=col, value=header)
    
    # 통합된 데이터 추가
    row_idx = 2
    for (도서영역별, 통합코드), group_data in grouped_data.items():
        ws2.cell(row=row_idx, column=1, value=도서영역별)
        ws2.cell(row=row_idx, column=2, value=통합코드)
        
        col_idx = 3
        for i, header in enumerate(headers):
            if header not in ['도서영역별', '도서코드', '도서', '통합코드', '통합여부']:
                ws2.cell(row=row_idx, column=col_idx, value=group_data.get(i, 0))
                col_idx += 1
        
        # 도서명(개정/통합) 추가
        도서명 = code_to_title.get(통합코드, "원본도서")
        ws2.cell(row=row_idx, column=col_idx, value=도서명)
        
        row_idx += 1
    
    # 6. 결과 저장
    new_workbook.save(output_file)
    print(f"✅ 결과 저장 완료: {output_file}")


# 실행 예시
input_file = "도서데이터.xlsx"   # 사용자가 입력한 파일
output_file = "도서데이터_통합.xlsx"
merge_sales_with_new_editions(input_file, output_file)
