import customtkinter as ctk
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
import os
from datetime import datetime
import threading

# CustomTkinter 설정
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

class SalesDataMergerGUI:
    def __init__(self):
        self.root = ctk.CTk()
        self.root.title("도서 판매 데이터 통합 시스템")
        self.root.geometry("900x700")
        self.root.minsize(800, 600)
        
        # 색상 테마 정의 (토스 스타일)
        self.colors = {
            'primary': '#0064FF',      # 토스 블루
            'secondary': '#00D4AA',    # 토스 민트
            'accent': '#FF6B6B',       # 토스 레드
            'success': '#00D4AA',      # 성공 그린
            'warning': '#FFB800',      # 경고 옐로우
            'background': '#F8F9FA',   # 배경 그레이
            'surface': '#FFFFFF',      # 카드 배경
            'text_primary': '#1A1A1A', # 주요 텍스트
            'text_secondary': '#6C757D' # 보조 텍스트
        }
        
        self.setup_ui()
        
    def setup_ui(self):
        # 메인 컨테이너
        main_frame = ctk.CTkFrame(self.root, fg_color="transparent")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # 제목 섹션
        self.create_title_section(main_frame)
        
        # 입력 섹션
        self.create_input_section(main_frame)
        
        # 옵션 섹션
        self.create_options_section(main_frame)
        
        # 실행 섹션
        self.create_execution_section(main_frame)
        
        # 로그 섹션
        self.create_log_section(main_frame)
        
    def create_title_section(self, parent):
        title_frame = ctk.CTkFrame(parent, fg_color=self.colors['surface'], corner_radius=15)
        title_frame.pack(fill="x", pady=(0, 20))
        
        title_label = ctk.CTkLabel(
            title_frame,
            text="📊 도서 판매 데이터 통합 시스템",
            font=ctk.CTkFont(family="SF Pro Display", size=24, weight="bold"),
            text_color=self.colors['text_primary']
        )
        title_label.pack(pady=20)
        
        subtitle_label = ctk.CTkLabel(
            title_frame,
            text="구판과 개정판의 판매 데이터를 통합하여 분석할 수 있습니다",
            font=ctk.CTkFont(family="SF Pro Text", size=14),
            text_color=self.colors['text_secondary']
        )
        subtitle_label.pack(pady=(0, 20))
        
    def create_input_section(self, parent):
        input_frame = ctk.CTkFrame(parent, fg_color=self.colors['surface'], corner_radius=15)
        input_frame.pack(fill="x", pady=(0, 20))
        
        # 섹션 제목
        section_title = ctk.CTkLabel(
            input_frame,
            text="📁 파일 선택",
            font=ctk.CTkFont(family="SF Pro Display", size=16, weight="bold"),
            text_color=self.colors['text_primary']
        )
        section_title.pack(pady=(20, 15), anchor="w", padx=20)
        
        # 입력 파일 선택
        input_file_frame = ctk.CTkFrame(input_frame, fg_color="transparent")
        input_file_frame.pack(fill="x", padx=20, pady=(0, 15))
        
        self.input_file_var = tk.StringVar()
        input_label = ctk.CTkLabel(
            input_file_frame,
            text="입력 파일:",
            font=ctk.CTkFont(family="SF Pro Text", size=12),
            text_color=self.colors['text_primary']
        )
        input_label.pack(anchor="w")
        
        input_file_row = ctk.CTkFrame(input_file_frame, fg_color="transparent")
        input_file_row.pack(fill="x", pady=(5, 0))
        
        self.input_file_entry = ctk.CTkEntry(
            input_file_row,
            textvariable=self.input_file_var,
            placeholder_text="Excel 파일을 선택하세요...",
            font=ctk.CTkFont(family="SF Pro Text", size=12),
            height=35
        )
        self.input_file_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        input_browse_btn = ctk.CTkButton(
            input_file_row,
            text="찾아보기",
            command=self.browse_input_file,
            font=ctk.CTkFont(family="SF Pro Text", size=12, weight="bold"),
            fg_color=self.colors['primary'],
            hover_color=self.colors['secondary'],
            height=35,
            width=80
        )
        input_browse_btn.pack(side="right")
        
        # 출력 파일 선택
        output_file_frame = ctk.CTkFrame(input_frame, fg_color="transparent")
        output_file_frame.pack(fill="x", padx=20, pady=(0, 20))
        
        self.output_file_var = tk.StringVar()
        output_label = ctk.CTkLabel(
            output_file_frame,
            text="출력 파일:",
            font=ctk.CTkFont(family="SF Pro Text", size=12),
            text_color=self.colors['text_primary']
        )
        output_label.pack(anchor="w")
        
        output_file_row = ctk.CTkFrame(output_file_frame, fg_color="transparent")
        output_file_row.pack(fill="x", pady=(5, 0))
        
        self.output_file_entry = ctk.CTkEntry(
            output_file_row,
            textvariable=self.output_file_var,
            placeholder_text="결과 파일 저장 위치를 선택하세요...",
            font=ctk.CTkFont(family="SF Pro Text", size=12),
            height=35
        )
        self.output_file_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        output_browse_btn = ctk.CTkButton(
            output_file_row,
            text="찾아보기",
            command=self.browse_output_file,
            font=ctk.CTkFont(family="SF Pro Text", size=12, weight="bold"),
            fg_color=self.colors['primary'],
            hover_color=self.colors['secondary'],
            height=35,
            width=80
        )
        output_browse_btn.pack(side="right")
        
    def create_options_section(self, parent):
        options_frame = ctk.CTkFrame(parent, fg_color=self.colors['surface'], corner_radius=15)
        options_frame.pack(fill="x", pady=(0, 20))
        
        # 섹션 제목
        section_title = ctk.CTkLabel(
            options_frame,
            text="⚙️ 처리 옵션",
            font=ctk.CTkFont(family="SF Pro Display", size=16, weight="bold"),
            text_color=self.colors['text_primary']
        )
        section_title.pack(pady=(20, 15), anchor="w", padx=20)
        
        # 옵션 체크박스들
        options_container = ctk.CTkFrame(options_frame, fg_color="transparent")
        options_container.pack(fill="x", padx=20, pady=(0, 20))
        
        self.backup_original = tk.BooleanVar(value=True)
        backup_checkbox = ctk.CTkCheckBox(
            options_container,
            text="원본 파일 백업 생성",
            variable=self.backup_original,
            font=ctk.CTkFont(family="SF Pro Text", size=12),
            fg_color=self.colors['primary'],
            hover_color=self.colors['secondary']
        )
        backup_checkbox.pack(anchor="w", pady=2)
        
        self.show_preview = tk.BooleanVar(value=True)
        preview_checkbox = ctk.CTkCheckBox(
            options_container,
            text="처리 전 미리보기 표시",
            variable=self.show_preview,
            font=ctk.CTkFont(family="SF Pro Text", size=12),
            fg_color=self.colors['primary'],
            hover_color=self.colors['secondary']
        )
        preview_checkbox.pack(anchor="w", pady=2)
        
        self.auto_open = tk.BooleanVar(value=False)
        auto_open_checkbox = ctk.CTkCheckBox(
            options_container,
            text="처리 완료 후 결과 파일 자동 열기",
            variable=self.auto_open,
            font=ctk.CTkFont(family="SF Pro Text", size=12),
            fg_color=self.colors['primary'],
            hover_color=self.colors['secondary']
        )
        auto_open_checkbox.pack(anchor="w", pady=2)
        
    def create_execution_section(self, parent):
        execution_frame = ctk.CTkFrame(parent, fg_color=self.colors['surface'], corner_radius=15)
        execution_frame.pack(fill="x", pady=(0, 20))
        
        # 버튼 컨테이너
        button_container = ctk.CTkFrame(execution_frame, fg_color="transparent")
        button_container.pack(pady=20)
        
        # 실행 버튼
        self.execute_btn = ctk.CTkButton(
            button_container,
            text="🚀 데이터 통합 실행",
            command=self.execute_merge,
            font=ctk.CTkFont(family="SF Pro Display", size=14, weight="bold"),
            fg_color=self.colors['success'],
            hover_color=self.colors['secondary'],
            height=45,
            width=200
        )
        self.execute_btn.pack(side="left", padx=(0, 10))
        
        # 미리보기 버튼
        preview_btn = ctk.CTkButton(
            button_container,
            text="👁️ 미리보기",
            command=self.preview_data,
            font=ctk.CTkFont(family="SF Pro Display", size=14, weight="bold"),
            fg_color=self.colors['primary'],
            hover_color=self.colors['secondary'],
            height=45,
            width=120
        )
        preview_btn.pack(side="left", padx=(0, 10))
        
        # 초기화 버튼
        reset_btn = ctk.CTkButton(
            button_container,
            text="🔄 초기화",
            command=self.reset_form,
            font=ctk.CTkFont(family="SF Pro Display", size=14, weight="bold"),
            fg_color=self.colors['warning'],
            hover_color=self.colors['accent'],
            height=45,
            width=100
        )
        reset_btn.pack(side="left")
        
    def create_log_section(self, parent):
        log_frame = ctk.CTkFrame(parent, fg_color=self.colors['surface'], corner_radius=15)
        log_frame.pack(fill="both", expand=True)
        
        # 섹션 제목
        section_title = ctk.CTkLabel(
            log_frame,
            text="📋 처리 로그",
            font=ctk.CTkFont(family="SF Pro Display", size=16, weight="bold"),
            text_color=self.colors['text_primary']
        )
        section_title.pack(pady=(20, 15), anchor="w", padx=20)
        
        # 로그 텍스트 영역
        self.log_text = ctk.CTkTextbox(
            log_frame,
            font=ctk.CTkFont(family="SF Mono", size=11),
            fg_color="#F8F9FA",
            text_color=self.colors['text_primary'],
            wrap="word"
        )
        self.log_text.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        # 로그 컨트롤 버튼
        log_controls = ctk.CTkFrame(log_frame, fg_color="transparent")
        log_controls.pack(fill="x", padx=20, pady=(0, 20))
        
        clear_log_btn = ctk.CTkButton(
            log_controls,
            text="로그 지우기",
            command=self.clear_log,
            font=ctk.CTkFont(family="SF Pro Text", size=11),
            fg_color=self.colors['text_secondary'],
            hover_color=self.colors['accent'],
            height=30,
            width=80
        )
        clear_log_btn.pack(side="left")
        
        save_log_btn = ctk.CTkButton(
            log_controls,
            text="로그 저장",
            command=self.save_log,
            font=ctk.CTkFont(family="SF Pro Text", size=11),
            fg_color=self.colors['text_secondary'],
            hover_color=self.colors['accent'],
            height=30,
            width=80
        )
        save_log_btn.pack(side="left", padx=(10, 0))
        
    def browse_input_file(self):
        filename = filedialog.askopenfilename(
            title="입력 파일 선택",
            filetypes=[
                ("Excel files", "*.xlsx *.xls"), 
                ("Excel 2007+ files", "*.xlsx"),
                ("Excel 97-2003 files", "*.xls"),
                ("All files", "*.*")
            ]
        )
        if filename:
            self.input_file_var.set(filename)
            # 자동으로 출력 파일명 생성
            base_name = os.path.splitext(os.path.basename(filename))[0]
            output_name = f"{base_name}_통합.xlsx"
            output_path = os.path.join(os.path.dirname(filename), output_name)
            self.output_file_var.set(output_path)
            self.log_message(f"입력 파일 선택: {filename}")
            
            # 파일 확장자 확인 및 안내
            file_ext = os.path.splitext(filename)[1].lower()
            if file_ext == '.xls':
                self.log_message("⚠️ .xls 파일 선택됨 - xlrd 패키지가 필요합니다.")
                self.log_message("   만약 오류가 발생하면 'pip install xlrd>=2.0.1'을 실행하세요.")
            
    def browse_output_file(self):
        filename = filedialog.asksaveasfilename(
            title="출력 파일 저장",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.output_file_var.set(filename)
            self.log_message(f"출력 파일 설정: {filename}")
            
    def log_message(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {message}\n"
        self.log_text.insert("end", log_entry)
        self.log_text.see("end")
        self.root.update_idletasks()
        
    def clear_log(self):
        self.log_text.delete("1.0", "end")
        
    def save_log(self):
        filename = filedialog.asksaveasfilename(
            title="로그 파일 저장",
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
        )
        if filename:
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(self.log_text.get("1.0", "end"))
            self.log_message(f"로그 파일 저장 완료: {filename}")
            
    def preview_data(self):
        input_file = self.input_file_var.get()
        if not input_file or not os.path.exists(input_file):
            messagebox.showerror("오류", "입력 파일을 선택해주세요.")
            return
            
        try:
            self.log_message("데이터 미리보기 시작...")
            
            # 파일 확장자 확인
            file_ext = os.path.splitext(input_file)[1].lower()
            if file_ext == '.xls':
                self.log_message("📋 .xls 파일 감지 - xlrd 엔진 사용")
            else:
                self.log_message("📋 .xlsx 파일 감지 - openpyxl 엔진 사용")
            
            # 시트 목록 확인
            excel_file = pd.ExcelFile(input_file)
            self.log_message(f"발견된 시트: {excel_file.sheet_names}")
            
            # 각 시트의 기본 정보 표시
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(input_file, sheet_name=sheet_name)
                self.log_message(f"시트 '{sheet_name}': {len(df)}행, {len(df.columns)}열")
                self.log_message(f"  컬럼: {list(df.columns)}")
                
        except ImportError as e:
            if "xlrd" in str(e):
                error_msg = "xlrd 패키지가 설치되지 않았습니다.\n\n해결 방법:\n1. 터미널에서 'pip install xlrd>=2.0.1' 실행\n2. 또는 .xlsx 파일로 변환 후 사용"
                self.log_message("❌ xlrd 패키지 누락 오류")
                messagebox.showerror("의존성 오류", error_msg)
            else:
                self.log_message(f"❌ 패키지 오류: {str(e)}")
                messagebox.showerror("패키지 오류", f"필요한 패키지가 설치되지 않았습니다:\n{str(e)}")
        except Exception as e:
            self.log_message(f"❌ 미리보기 오류: {str(e)}")
            messagebox.showerror("오류", f"미리보기 중 오류가 발생했습니다:\n{str(e)}")
            
    def reset_form(self):
        self.input_file_var.set("")
        self.output_file_var.set("")
        self.backup_original.set(True)
        self.show_preview.set(True)
        self.auto_open.set(False)
        self.log_message("폼이 초기화되었습니다.")
        
    def execute_merge(self):
        input_file = self.input_file_var.get()
        output_file = self.output_file_var.get()
        
        if not input_file or not output_file:
            messagebox.showerror("오류", "입력 파일과 출력 파일을 모두 선택해주세요.")
            return
            
        if not os.path.exists(input_file):
            messagebox.showerror("오류", "입력 파일이 존재하지 않습니다.")
            return
            
        # 백그라운드에서 실행
        self.execute_btn.configure(state="disabled", text="처리 중...")
        thread = threading.Thread(target=self._execute_merge_thread, args=(input_file, output_file))
        thread.daemon = True
        thread.start()
        
    def _execute_merge_thread(self, input_file, output_file):
        try:
            self.log_message("데이터 통합 작업 시작...")
            
            # 백업 생성
            if self.backup_original.get():
                backup_file = f"{os.path.splitext(output_file)[0]}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                import shutil
                shutil.copy2(input_file, backup_file)
                self.log_message(f"원본 파일 백업 생성: {backup_file}")
            
            # 데이터 통합 실행
            self.merge_sales_with_new_editions(input_file, output_file)
            
            self.log_message("✅ 데이터 통합 완료!")
            
            # 자동 열기
            if self.auto_open.get():
                import subprocess
                try:
                    subprocess.Popen(['start', output_file], shell=True)
                    self.log_message("결과 파일을 자동으로 열었습니다.")
                except:
                    self.log_message("결과 파일 자동 열기 실패")
                    
            messagebox.showinfo("완료", "데이터 통합이 성공적으로 완료되었습니다!")
            
        except Exception as e:
            error_msg = f"처리 중 오류가 발생했습니다:\n{str(e)}"
            self.log_message(f"❌ 오류: {str(e)}")
            messagebox.showerror("오류", error_msg)
            
        finally:
            self.execute_btn.configure(state="normal", text="🚀 데이터 통합 실행")
            
    def merge_sales_with_new_editions(self, input_file, output_file):
        self.log_message("1. 엑셀 파일에서 시트 불러오는 중...")
        
        # 파일 확장자 확인
        file_ext = os.path.splitext(input_file)[1].lower()
        if file_ext == '.xls':
            self.log_message("📋 .xls 파일 감지 - xlrd 엔진 사용")
        else:
            self.log_message("📋 .xlsx 파일 감지 - openpyxl 엔진 사용")
        
        try:
            # 1. 시트 불러오기
            sales_df = pd.read_excel(input_file, sheet_name="도서판매추이")
            mapping_df = pd.read_excel(input_file, sheet_name="개정판 정보")
            
            self.log_message(f"  - 도서판매추이: {len(sales_df)}행")
            self.log_message(f"  - 개정판 정보: {len(mapping_df)}행")
        except ImportError as e:
            if "xlrd" in str(e):
                raise ImportError("xlrd 패키지가 설치되지 않았습니다. 'pip install xlrd>=2.0.1'을 실행하세요.")
            else:
                raise ImportError(f"필요한 패키지가 설치되지 않았습니다: {str(e)}")
        except Exception as e:
            raise Exception(f"엑셀 파일 읽기 오류: {str(e)}")

        # 2. 매핑 dict (구판 → 개정판)
        self.log_message("2. 매핑 정보 생성 중...")
        map_dict = dict(zip(mapping_df['x_code'], mapping_df['y_code']))
        self.log_message(f"  - 매핑된 도서: {len(map_dict)}개")

        # 3. "통합코드" 생성 (개정판 정보가 있으면 개정판 코드, 없으면 원래 코드)
        self.log_message("3. 통합코드 컬럼 생성 중...")
        # 개정판 정보가 있는 도서는 개정판 코드로, 없는 도서는 원래 코드 유지
        sales_df['통합코드'] = sales_df['도서코드'].apply(
            lambda x: map_dict.get(x, x)  # 매핑에 있으면 개정판 코드, 없으면 원래 코드
        )

        # 4. 통합여부 표시 (통합된 도서는 "통합-구판제목" 형태로)
        self.log_message("4. 통합여부 표시 중...")
        sales_df['통합여부'] = sales_df.apply(
            lambda r: f"통합-{r['도서']}" if r['도서코드'] in map_dict else "단일", axis=1
        )
        
        integrated_count = len(sales_df[sales_df['통합여부'] == "통합"])
        self.log_message(f"  - 통합 대상: {integrated_count}개")

        # 5. 구판/개정판 합산 (도서영역별+통합코드 기준으로 일자별 합산)
        self.log_message("5. 판매 데이터 통합 중...")
        id_cols = ['도서영역별', '통합코드']
        value_cols = [c for c in sales_df.columns if c not in ['도서영역별','도서코드','도서','통합코드','통합여부']]
        merged_df = sales_df.groupby(id_cols)[value_cols].sum().reset_index()
        
        self.log_message(f"  - 통합 후: {len(merged_df)}개 도서")

        # 6. 도서명 처리 및 통합여부 추가
        self.log_message("6. 도서명 정보 및 통합여부 추가 중...")
        
        # 구판 코드 → 구판 제목 매핑
        old_code_to_title = dict(zip(mapping_df['x_code'], mapping_df['x_title']))
        
        # 개정판 코드 → 개정판 제목 매핑
        new_code_to_title = dict(zip(mapping_df['y_code'], mapping_df['y_title']))
        
        # 원본 도서명 정보 가져오기 (통합코드별로 첫 번째 원본 도서명 사용)
        original_titles = sales_df.groupby('통합코드')['도서'].first().to_dict()
        
        # 도서명 처리: 통합코드가 개정판 코드인 경우 개정판 제목, 아니면 원래 도서명
        merged_df['도서명'] = merged_df['통합코드'].map(new_code_to_title)
        merged_df['도서명'] = merged_df['도서명'].fillna(merged_df['통합코드'].map(original_titles))
        
        # 통합여부 추가: 통합된 도서는 "통합-구판제목" 형태로 표시
        # 통합코드가 개정판 매핑에 있으면 해당 구판의 제목을 찾아서 표시
        def get_integration_status(row):
            integrated_code = row['통합코드']
            # 통합코드가 개정판 코드인 경우 (즉, 구판이 개정판으로 통합된 경우)
            if integrated_code in new_code_to_title:
                # 해당 개정판 코드에 매핑된 구판 코드들을 찾기
                related_old_codes = [old_code for old_code, new_code in map_dict.items() if new_code == integrated_code]
                if related_old_codes:
                    # 첫 번째 구판 제목 사용
                    old_title = old_code_to_title.get(related_old_codes[0], '알수없음')
                    return f"통합-{old_title}"
            return "단일"
        
        merged_df['통합여부'] = merged_df.apply(get_integration_status, axis=1)
        
        # 컬럼 순서 재배치: 도서영역별, 통합코드, 도서명, 통합여부, 나머지 컬럼들
        self.log_message("7. 컬럼 순서 재배치 중...")
        first_cols = ['도서영역별', '통합코드', '도서명', '통합여부']
        remaining_cols = [col for col in merged_df.columns if col not in first_cols]
        merged_df = merged_df[first_cols + remaining_cols]

        # 8. 저장 (서식 제거, 도서판매추이 구조 유지)
        self.log_message("8. 결과 파일 저장 중...")
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            merged_df.to_excel(writer, sheet_name="도서판매추이(통합)", index=False)
            
            # 워크시트 가져오기
            worksheet = writer.sheets["도서판매추이(통합)"]
            
            # 모든 셀의 서식 제거 (안전한 방법)
            from openpyxl.styles import Font, Border, PatternFill, Alignment
            
            for row in worksheet.iter_rows():
                for cell in row:
                    # 기본 서식으로 설정
                    cell.font = Font()
                    cell.border = Border()
                    cell.fill = PatternFill(fill_type=None)
                    cell.number_format = 'General'
                    cell.alignment = Alignment()

        self.log_message(f"✅ 결과 저장 완료: {output_file}")
        
    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    app = SalesDataMergerGUI()
    app.run()
