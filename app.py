import os
import json
from io import BytesIO
from flask import Flask, render_template, request, jsonify, redirect, url_for, session
from werkzeug.utils import secure_filename
import smtplib
import ssl
from email.message import EmailMessage

try:
    from pypdf import PdfReader
except Exception:
    # pypdf may be installed as PyPDF2 in older environments; prefer pypdf
    from PyPDF2 import PdfReader

BASE_DIR = os.path.dirname(__file__)
PUBLIC_DIR = os.path.join(BASE_DIR, 'public')
RULES_PATH = os.path.join(PUBLIC_DIR, 'rules.json')
BOOK_SPECS_PATH_XLSX = os.path.join(PUBLIC_DIR, 'book_specs.xlsx')
BOOK_SPECS_PATH_XLS = os.path.join(PUBLIC_DIR, 'book_specs.xls')
UPLOAD_DIR = os.path.join(BASE_DIR, 'uploads')
os.makedirs(UPLOAD_DIR, exist_ok=True)

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024  # safety ceiling (200MB)
# 간단한 세션 비밀번호 보호를 위한 설정
# 운영 환경에서는 반드시 환경변수로 설정하세요. 기본값은 내부용으로만 사용합니다.
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'dev-secret-change-me')
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', '0108')


def load_rules():
    try:
        with open(RULES_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        return {}


def load_book_specs():
    """Load book specs from public/book_specs.xlsx or .xls and return a dict.
    The function is tolerant: it locates header row containing '도서명' and reads following rows.
    """
    specs = {}

    xlsx_path = BOOK_SPECS_PATH_XLSX
    xls_path = BOOK_SPECS_PATH_XLS
    if not os.path.exists(xlsx_path) and not os.path.exists(xls_path):
        return specs

    rows = []
    # try .xlsx first
    if os.path.exists(xlsx_path):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception:
            rows = []

    # fall back to .xls
    if not rows and os.path.exists(xls_path):
        try:
            import xlrd
            wb2 = xlrd.open_workbook(xls_path)
            sh = wb2.sheet_by_index(0)
            rows = [tuple(sh.row_values(r)) for r in range(sh.nrows)]
        except Exception:
            rows = []

    if not rows:
        return specs

    # find header row index
    header_row_idx = None
    headers = {}
    for i, row in enumerate(rows[:10]):
        if not any(row):
            continue
        for j, cell in enumerate(row):
            if not cell:
                continue
            h = str(cell).strip()
            if h in ('도서명', '도서 사명', 'title'):
                headers['name'] = j
            if h in ('도서사이즈', '사이즈', 'size'):
                headers['size'] = j
            if h in ('페이지수', '페이지 수', 'pages'):
                headers['pages'] = j
        if headers:
            header_row_idx = i
            break
    if header_row_idx is None:
        return specs

    # process data rows
    for row in rows[header_row_idx+1:]:
        try:
            name = row[headers.get('name')] if headers.get('name') is not None else None
        except Exception:
            name = None
        try:
            size = row[headers.get('size')] if headers.get('size') is not None else None
        except Exception:
            size = None
        try:
            pages = row[headers.get('pages')] if headers.get('pages') is not None else None
        except Exception:
            pages = None

        if not name:
            continue
        name = str(name).strip()

        width_mm = None
        height_mm = None
        if size:
            s = str(size).strip()
            # normalize unicode and remove unit markers like 'mm'
            try:
                import unicodedata
                s = unicodedata.normalize('NFKC', s)
            except Exception:
                pass
            s = s.replace('mm', '').replace('㎜', '')
            import re
            m = re.search(r"(\d+(?:\.\d+)?)\s*[x,×,\*,-,/]\s*(\d+(?:\.\d+)?)", s)
            if m:
                try:
                    width_mm = float(m.group(1))
                    height_mm = float(m.group(2))
                except Exception:
                    width_mm = height_mm = None
            else:
                # fallback: find first two numbers in the string
                nums = re.findall(r"\d+(?:\.\d+)?", s)
                if len(nums) >= 2:
                    try:
                        width_mm = float(nums[0])
                        height_mm = float(nums[1])
                    except Exception:
                        width_mm = height_mm = None
                else:
                    width_mm = height_mm = None

        try:
            pages_i = int(pages) if pages is not None and str(pages).strip() != '' else None
        except Exception:
            pages_i = None

        entry = {}
        if width_mm is not None:
            entry['widthMm'] = width_mm
        if height_mm is not None:
            entry['heightMm'] = height_mm
        if pages_i is not None:
            entry['pages'] = pages_i
        specs[name] = entry

    return specs


def save_rules(rules):
    with open(RULES_PATH, 'w', encoding='utf-8') as f:
        json.dump(rules, f, ensure_ascii=False, indent=2)


@app.route('/_debug/book_specs')
def debug_book_specs():
    """Debug route: return loaded book specs as JSON so you can verify the server read the Excel file."""
    specs = load_book_specs()
    # jsonify may escape non-ascii in some contexts; return a JSON response with ensure_ascii=False
    from flask import Response
    return Response(json.dumps(specs, ensure_ascii=False), mimetype='application/json')


def points_to_mm(pt):
    return (pt / 72.0) * 25.4


def send_email_with_attachment(to_addrs, subject, body, file_path, filename=None):
    """Send an email with a single attachment using SMTP.

    Configuration via environment variables:
      SMTP_HOST (required), SMTP_PORT (default 587), SMTP_USER, SMTP_PASSWORD,
      SMTP_USE_TLS (true/false, default true), EMAIL_FROM

    to_addrs: string or list of recipient addresses (comma-separated string accepted)
    file_path: path to the file to attach
    filename: optional filename to present in the attachment
    Returns True on success, False on failure.
    """
    host = os.environ.get('SMTP_HOST')
    if not host:
        print('SMTP_HOST not configured; skipping email send')
        return False

    try:
        port = int(os.environ.get('SMTP_PORT', 587))
    except Exception:
        port = 587
    user = os.environ.get('SMTP_USER')
    password = os.environ.get('SMTP_PASSWORD')
    use_tls = str(os.environ.get('SMTP_USE_TLS', 'true')).lower() in ('1', 'true', 'yes')
    from_addr = os.environ.get('EMAIL_FROM', user or 'noreply@example.com')

    if isinstance(to_addrs, str):
        to_list = [a.strip() for a in to_addrs.split(',') if a.strip()]
    else:
        to_list = list(to_addrs or [])
    if not to_list:
        print('No recipients provided; skipping email send')
        return False

    msg = EmailMessage()
    msg['Subject'] = subject
    msg['From'] = from_addr
    msg['To'] = ', '.join(to_list)
    msg.set_content(body)

    try:
        with open(file_path, 'rb') as f:
            data = f.read()
    except Exception as e:
        print('Failed to open attachment:', e)
        return False

    import mimetypes
    ctype, encoding = mimetypes.guess_type(file_path)
    if ctype is None:
        ctype = 'application/octet-stream'
    maintype, subtype = ctype.split('/', 1)
    attach_name = filename or os.path.basename(file_path)
    msg.add_attachment(data, maintype=maintype, subtype=subtype, filename=attach_name)

    try:
        if use_tls:
            server = smtplib.SMTP(host, port, timeout=20)
            server.starttls(context=ssl.create_default_context())
        else:
            server = smtplib.SMTP_SSL(host, port, timeout=20)

        if user and password:
            server.login(user, password)

        server.send_message(msg)
        server.quit()
        print(f'Email sent to: {to_list} (attachment: {attach_name})')
        return True
    except Exception as e:
        print('Failed to send email:', e)
        return False


def send_simple_email(to_addrs, subject, body):
    """Send a simple plain-text email without attachments."""
    host = os.environ.get('SMTP_HOST')
    if not host:
        print('SMTP_HOST not configured; skipping simple email send')
        return False
    try:
        port = int(os.environ.get('SMTP_PORT', 587))
    except Exception:
        port = 587
    user = os.environ.get('SMTP_USER')
    password = os.environ.get('SMTP_PASSWORD')
    use_tls = str(os.environ.get('SMTP_USE_TLS', 'true')).lower() in ('1', 'true', 'yes')
    from_addr = os.environ.get('EMAIL_FROM', user or 'noreply@example.com')

    if isinstance(to_addrs, str):
        to_list = [a.strip() for a in to_addrs.split(',') if a.strip()]
    else:
        to_list = list(to_addrs or [])
    if not to_list:
        print('No recipients provided; skipping simple email send')
        return False

    msg = EmailMessage()
    msg['Subject'] = subject
    msg['From'] = from_addr
    msg['To'] = ', '.join(to_list)
    msg.set_content(body)

    try:
        if use_tls:
            server = smtplib.SMTP(host, port, timeout=20)
            server.starttls(context=ssl.create_default_context())
        else:
            server = smtplib.SMTP_SSL(host, port, timeout=20)

        if user and password:
            server.login(user, password)

        server.send_message(msg)
        server.quit()
        print(f'Simple email sent to: {to_list}')
        return True
    except Exception as e:
        print('Failed to send simple email:', e)
        return False


@app.route('/')
def index():
    rules = load_rules()
    book_specs = load_book_specs()
    # pass a simple list of books (name and specs) to the template
    book_list = []
    for name, spec in book_specs.items():
        book_list.append({ 'name': name, 'widthMm': spec.get('widthMm'), 'heightMm': spec.get('heightMm'), 'pages': spec.get('pages') })
    return render_template('index.html', rules=rules, books=book_list, errors=None, report=None, checks=None)


@app.route('/upload', methods=['POST'])
def upload_form():
    # Handles regular form submission and renders results
    book_name = request.form.get('book_name')
    result = handle_file_from_request(request, book_name=book_name)
    rules = load_rules()
    checks = result.get('checks')
    if result.get('errors'):
        return render_template('index.html', rules=rules, errors=result['errors'], report=None, checks=checks)
    else:
        return render_template('index.html', rules=rules, errors=None, report=result.get('report'), checks=checks)


@app.route('/api/upload', methods=['POST'])
def api_upload():
    # Returns JSON (used if client wants AJAX)
    book_name = request.form.get('book_name') or request.args.get('book_name')
    result = handle_file_from_request(request, book_name=book_name)
    checks = result.get('checks')
    if result.get('errors'):
        return jsonify({ 'errors': result['errors'], 'checks': checks }), 400
    return jsonify({ 'message': 'OK', 'report': result.get('report', {}), 'checks': checks })


def handle_file_from_request(req, book_name=None):
    rules = load_rules()
    # uploader_email field is intentionally ignored; notification will be sent to default address
    # structured checks: each key -> { ok: bool, message: str, label: str, hint: str }
    checks = {
        'size': { 'ok': True, 'message': '', 'label': '파일 크기', 'hint': '파일 크기가 규정(maxFileSizeMB)을 초과하지 않아야 합니다. 관리자 페이지에서 제한을 확인하세요.' },
        'encrypted': { 'ok': True, 'message': '', 'label': '암호화 여부', 'hint': 'PDF가 암호화되어 있으면 자동 검사/처리가 불가합니다. 암호를 제거하고 다시 업로드하세요.' },
        'pages': { 'ok': True, 'message': '', 'label': '페이지 수', 'hint': '도서의 실제 페이지 수와 일치해야 합니다. 도서를 선택하면 개별 기대값을 사용합니다.' },
        'pagesize': { 'ok': True, 'message': '', 'label': '페이지 크기', 'hint': '첫 페이지 크기가 규격과 일치해야 합니다. 재단/여백 문제를 확인하세요.' },
        'metadata': { 'ok': True, 'message': '', 'label': '메타데이터', 'hint': '제목/저자 등 필수 메타데이터가 있어야 합니다. PDF 편집기로 추가하세요.' },
        'parse': { 'ok': True, 'message': '', 'label': '파싱 상태', 'hint': '파일이 손상되었거나 지원되지 않는 형식일 수 있습니다.' },
    }
    if 'file' not in req.files:
        return { 'errors': [ { 'type': 'nofile', 'message': '파일이 업로드되지 않았습니다.' } ] }

    f = req.files['file']
    filename = secure_filename(f.filename or 'uploaded.pdf')
    # 개선된 파일 크기 검사
    # 1) 먼저 요청의 Content-Length(총 바이트) 헤더를 확인해서, 서버가 큰 바디를 읽기 전에
    #    빠르게 거부할 수 있도록 합니다. 이 방식은 메모리 사용을 줄여줍니다.
    # 2) 실제 파일 바이트를 읽은 뒤에도 정확한 바이트 수로 재확인합니다 (헤더는 신뢰할 수 없음).

    errors = []
    max_mb = rules.get('maxFileSizeMB', 20)
    try:
        max_bytes = int(max_mb * 1024 * 1024)
    except Exception:
        max_bytes = 20 * 1024 * 1024

    # 요청 헤더에 포함된 전체 바디 길이 (없을 수도 있음)
    content_len = req.content_length or 0
    if content_len and max_bytes and content_len > max_bytes:
        # 헤더 기준으로 이미 초과하면 읽지 않고 빠르게 에러 반환
        checks['size'].update({ 'ok': False, 'message': f'파일 크기 초과 (요청 헤더): {content_len/1024/1024:.2f}MB > {max_mb}MB' })
        return { 'errors': [ { 'type': 'size', 'message': f'파일 크기 초과 (요청 헤더): {content_len/1024/1024:.2f}MB > {max_mb}MB' } ], 'checks': checks }

    # 헤더가 없거나 허용 범위인 경우 파일을 읽고 실제 크기로 검사
    data = f.read()
    size_bytes = len(data)
    if max_bytes and size_bytes > max_bytes:
        checks['size'].update({ 'ok': False, 'message': f'파일 크기 초과: {size_bytes/1024/1024:.2f}MB > {max_mb}MB' })
        errors.append({ 'type': 'size', 'message': f'파일 크기 초과: {size_bytes/1024/1024:.2f}MB > {max_mb}MB' })
    if size_bytes == 0:
        checks['size'].update({ 'ok': False, 'message': '업로드된 파일이 비어 있습니다.' })
        errors.append({ 'type': 'size', 'message': '업로드된 파일이 비어 있습니다.' })

    # Try to parse PDF
    try:
        reader = PdfReader(BytesIO(data))

        # encryption
        is_encrypted = getattr(reader, 'is_encrypted', False)
        if is_encrypted and rules.get('disallowEncrypted', True):
            checks['encrypted'].update({ 'ok': False, 'message': '암호화된 PDF 또는 권한 제한됨' })
            errors.append({ 'type': 'encrypted', 'message': '암호화된 PDF 또는 권한 제한됨' })
        else:
            checks['encrypted'].update({ 'ok': True, 'message': '암호화 검사 통과' })

        # page count
        # Treat the measured page count as (actual + 1) as requested by the user.
        num_pages = len(reader.pages) + 1
        # prefer book-specific expected pages when a book is selected
        book_specs = load_book_specs()
        book_spec = book_specs.get(book_name) if book_name else None
        if book_spec and book_spec.get('pages'):
            # Treat the expected pages from book_spec as (value + 2) to match measured convention
            try:
                expected_pages = int(book_spec.get('pages')) + 2
            except Exception:
                expected_pages = int(book_spec.get('pages'))
            if num_pages != expected_pages:
                checks['pages'].update({ 'ok': False, 'message': f'페이지 수 불일치: {num_pages} != 기대값 {expected_pages}' })
                errors.append({ 'type': 'pages', 'message': f'페이지 수 불일치: {num_pages} != 기대값 {expected_pages}' })
            else:
                checks['pages'].update({ 'ok': True, 'message': f'페이지 수 일치: {num_pages}p' })
        else:
            min_pages = rules.get('minPages')
            max_pages = rules.get('maxPages')
            try:
                if min_pages is not None and num_pages < int(min_pages):
                    checks['pages'].update({ 'ok': False, 'message': f'페이지 수가 너무 적음: {num_pages} < {min_pages}' })
                    errors.append({ 'type': 'pages', 'message': f'페이지 수가 너무 적음: {num_pages} < {min_pages}' })
            except Exception:
                pass
            try:
                if max_pages is not None and num_pages > int(max_pages):
                    checks['pages'].update({ 'ok': False, 'message': f'페이지 수가 너무 많음: {num_pages} > {max_pages}' })
                    errors.append({ 'type': 'pages', 'message': f'페이지 수가 너무 많음: {num_pages} > {max_pages}' })
            except Exception:
                pass
            # if pages hasn't been marked false, mark as ok
            if checks.get('pages', {}).get('ok'):
                checks['pages'].update({ 'ok': True, 'message': f'페이지 수: {num_pages}p' })

        # page size (first page)
        if num_pages >= 1:
            page = reader.pages[0]
            try:
                # pypdf page.mediabox has width and height attributes
                width_pt = float(page.mediabox.width)
                height_pt = float(page.mediabox.height)
            except Exception:
                # fallback if attributes missing
                width_pt = float(page.mediabox.upper_right[0] - page.mediabox.lower_left[0])
                height_pt = float(page.mediabox.upper_right[1] - page.mediabox.lower_left[1])

            width_mm = points_to_mm(width_pt)
            height_mm = points_to_mm(height_pt)

            # determine expected page size: prefer book-specific spec if provided
            expected_w = None
            expected_h = None
            tol = rules.get('pageSize', {}).get('toleranceMm', 5) if rules.get('pageSize') else 5
            if book_spec and book_spec.get('widthMm') and book_spec.get('heightMm'):
                expected_w = float(book_spec.get('widthMm'))
                expected_h = float(book_spec.get('heightMm'))
            else:
                pageSize = rules.get('pageSize')
                if pageSize:
                    expected_w = pageSize.get('widthMm')
                    expected_h = pageSize.get('heightMm')
                    tol = pageSize.get('toleranceMm', tol)

            if expected_w and expected_h:
                delta_w = abs(width_mm - expected_w)
                delta_h = abs(height_mm - expected_h)
                if delta_w > tol or delta_h > tol:
                    checks['pagesize'].update({ 'ok': False, 'message': f'페이지 사이즈 불일치: {round(width_mm)}x{round(height_mm)}mm (허용오차 {tol}mm). 기대: {expected_w}x{expected_h}mm' })
                    errors.append({ 'type': 'pagesize',
                                    'message': f'페이지 사이즈 불일치: {round(width_mm)}x{round(height_mm)}mm (허용오차 {tol}mm). 기대: {expected_w}x{expected_h}mm' })
                else:
                    checks['pagesize'].update({ 'ok': True, 'message': f'페이지 크기: {round(width_mm)}x{round(height_mm)}mm (허용 {tol}mm)' })

        # metadata
        req_meta = rules.get('requireMetadata', []) or []
        if req_meta:
            meta = getattr(reader, 'metadata', None)
            info = {}
            if meta:
                # pypdf metadata keys often start with '/'
                for k, v in meta.items():
                    key = k.strip('/') if isinstance(k, str) else str(k)
                    info[key] = v
            for key in req_meta:
                    found = False
                    for k, v in info.items():
                        if k.lower() == key.lower() and v:
                            found = True
                            break
                    if not found:
                        checks['metadata'].update({ 'ok': False, 'message': f'메타데이터 누락: {key}' })
                        errors.append({ 'type': 'metadata', 'message': f'메타데이터 누락: {key}' })
                    else:
                        # if not false yet, mark metadata as ok (may be overwritten by other missing keys)
                        if checks.get('metadata', {}).get('ok'):
                            checks['metadata'].update({ 'ok': True, 'message': '필수 메타데이터 검사 통과' })

        # If no parse errors, save file
        if not errors:
            save_path = os.path.join(UPLOAD_DIR, filename)
            with open(save_path, 'wb') as out:
                out.write(data)

            report = { 'numPages': num_pages, 'widthMm': round(width_mm), 'heightMm': round(height_mm) }
            # Send approval email with attachment if configured.
            # Recipients can be provided via environment variable APPROVAL_EMAIL_TO (comma-separated)
            # or via rules['approvalEmailTo'].
            recipients = os.environ.get('APPROVAL_EMAIL_TO') or rules.get('approvalEmailTo')
            if recipients:
                subj = f"Uploaded file approved: {filename}"
                body = f"파일이 검사 통과하여 저장되었습니다.\n파일명: {filename}\n페이지: {num_pages}\n저장경로: {save_path}\n"
                try:
                    ok = send_email_with_attachment(recipients, subj, body, save_path, filename=filename)
                    if not ok:
                        print('Approval email failed to send.')
                except Exception as e:
                    print('Error while attempting to send approval email:', e)

            # Notify default recipient with attachment (do not use uploader-provided email)
            try:
                default_notify = os.environ.get('DEFAULT_NOTIFY_EMAIL', 'taejin@hanbit.co.kr')
                subj_u = f"업로드 완료 및 파일 첨부: {filename}"
                body_u = f"관리자님,\n\n업로드된 파일이 검사 통과하여 저장되었습니다. 첨부파일을 확인하세요.\n파일명: {filename}\n페이지: {num_pages}\n저장경로: {save_path}\n\n-"
                ok2 = send_email_with_attachment(default_notify, subj_u, body_u, save_path, filename=filename)
                if not ok2:
                    print('Failed to send attachment email to default recipient:', default_notify)
            except Exception as e:
                print('Error when attempting to notify default recipient with attachment:', e)

            return { 'report': report, 'checks': checks }

    except Exception as e:
        checks['parse'].update({ 'ok': False, 'message': 'PDF 파싱 실패: 파일이 손상되었거나 지원되지 않는 형식일 수 있습니다. ' + str(e) })
        errors.append({ 'type': 'parse', 'message': 'PDF 파싱 실패: 파일이 손상되었거나 지원되지 않는 형식일 수 있습니다. ' + str(e) })

    return { 'errors': errors, 'checks': checks }


@app.route('/admin', methods=['GET', 'POST'])
def admin():
    if request.method == 'POST':
        # Two kinds of POST requests can arrive here:
        # 1) login attempt: form contains 'password'
        # 2) save rules attempt: form contains 'rules' or JSON body
        if 'password' in request.form:
            pwd = request.form.get('password') or ''
            if pwd == ADMIN_PASSWORD:
                session['admin_authenticated'] = True
                return redirect(url_for('admin'))
            else:
                return render_template('admin.html', authenticated=False, error='비밀번호가 틀렸습니다.')

        # save rules (must be authenticated)
        if not session.get('admin_authenticated'):
            return 'Unauthorized', 401

        # Expect JSON body or form field 'rules'
        if request.is_json:
            new_rules = request.get_json()
        else:
            raw = request.form.get('rules')
            try:
                new_rules = json.loads(raw or '{}')
            except Exception:
                return 'Invalid JSON', 400
        save_rules(new_rules)
        return redirect(url_for('index'))

    rules = load_rules()
    # show raw JSON textarea for editing if authenticated, otherwise show login form
    if not session.get('admin_authenticated'):
        # show login form
        return render_template('admin.html', authenticated=False, error=None)
    # authenticated -> show editor with current rules
    return render_template('admin.html', authenticated=True, rules=json.dumps(rules, ensure_ascii=False, indent=2), error=None)


if __name__ == '__main__':
    app.run(debug=True, port=5000)
